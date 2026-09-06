import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from database import get_conn, get_credits_for_month, get_fixed_for_month

MONTHS_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")


def _style_header_row(ws, row_idx=1):
    for cell in ws[row_idx]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = length + 3


def generate_yearly_report(year):
    """
    Builds an in-memory .xlsx workbook:
      - 12 monthly detail sheets (income/credits/fixed/other + totals)
      - 1 summary sheet with a bar chart comparing months
    Returns a BytesIO (with .name set) ready for bot.send_document().
    """
    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "Жылдық қорытынды"

    summary_ws.append(["Ай", "Кирис", "Кредитлер", "Тұрақлы харажатлар",
                        "Басқа харажатлар", "Жалпы харажат", "Қалды"])
    _style_header_row(summary_ws)

    for month_num in range(1, 13):
        date_filter = f"{year}-{month_num:02d}"

        conn = get_conn()
        c = conn.cursor()
        c.execute("SELECT COALESCE(SUM(amount),0) FROM budget WHERE created_at LIKE %s",
                  (f"{date_filter}%",))
        income = float(c.fetchone()[0])

        c.execute("SELECT category, COALESCE(SUM(amount),0) FROM other_expenses WHERE created_at LIKE %s GROUP BY category",
                  (f"{date_filter}%",))
        other_by_cat = c.fetchall()
        conn.close()

        other_total = sum(float(a) for _, a in other_by_cat)

        credits = get_credits_for_month(date_filter)
        fixed = get_fixed_for_month(date_filter)
        credit_total = sum(float(a) for _, _, a, _ in credits)
        fixed_total = sum(float(a) for _, _, a, _ in fixed)
        total_expense = credit_total + fixed_total + other_total
        remaining = income - total_expense

        month_name = MONTHS_RU[month_num]
        summary_ws.append([month_name, income, credit_total, fixed_total,
                            other_total, total_expense, remaining])

        # ---- Per-month detail sheet ----
        ws = wb.create_sheet(title=month_name)
        ws.append(["Түри", "Аты/Категория", "Сумма"])
        _style_header_row(ws)

        if credits:
            for cid, name, amount, pay_day in credits:
                ws.append(["Кредит", name, float(amount)])
        if fixed:
            for fid, name, amount, pay_day in fixed:
                ws.append(["Тұрақлы харажат", name, float(amount)])
        if other_by_cat:
            for cat, amount in other_by_cat:
                ws.append(["Басқа харажат", cat, float(amount)])

        ws.append([])
        ws.append(["Кирис", "", income])
        ws.append(["Жалпы харажат", "", total_expense])
        ws.append(["Қалды", "", remaining])
        _autosize_columns(ws)

    # ---- Bar chart on the summary sheet ----
    chart = BarChart()
    chart.type = "col"
    chart.title = f"{year} — Айлар бойынша кирис ҳәм харажат"
    chart.y_axis.title = "Сум"
    chart.x_axis.title = "Ай"
    chart.width = 26
    chart.height = 13

    data = Reference(summary_ws, min_col=2, max_col=6, min_row=1, max_row=13)
    cats = Reference(summary_ws, min_col=1, min_row=2, max_row=13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    summary_ws.add_chart(chart, "I2")

    _autosize_columns(summary_ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = f"jyldyq_esap_{year}.xlsx"
    return buf

